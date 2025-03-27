import os
import pcbnew
import subprocess
import openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Define project paths
PROJECT_PATH = os.getcwd()
OUTPUT_PATH = os.path.join(PROJECT_PATH, "Outputs")
BOM_PATH = os.path.join(OUTPUT_PATH, "BOM")
PDF_PATH = os.path.join(OUTPUT_PATH, "PDF")
GERBER_PATH = os.path.join(OUTPUT_PATH, "GERBER/")
STEP_PATH = os.path.join(OUTPUT_PATH, "3D")

project_files = [f for f in os.listdir(PROJECT_PATH) if f.endswith(".kicad_pro")]
if project_files:
    PROJECT_NAME = os.path.splitext(project_files[0])[0]
else:
    raise FileNotFoundError("No KiCad project file (.kicad_pro) found in the current directory.")

# Ensure output directories exist
os.makedirs(BOM_PATH, exist_ok=True)
os.makedirs(PDF_PATH, exist_ok=True)
os.makedirs(GERBER_PATH, exist_ok=True)
os.makedirs(STEP_PATH, exist_ok=True)

# Paths to KiCad files
PCB_FILE = os.path.join(PROJECT_PATH, f"{PROJECT_NAME}.kicad_pcb")
SCH_FILE = os.path.join(PROJECT_PATH, f"{PROJECT_NAME}.kicad_sch")

# Function to execute KiCad CLI commands
def run_kicad_command(command, description):
    print(f"Executing: {description}...")
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    
    if result.returncode == 0:
        print(f"{description} completed successfully.")
    else:
        print(f"Error during {description}: {result.stderr}")

# Load PCB
pcb = pcbnew.LoadBoard(PCB_FILE)

# Dictionary to group components
bom_dict = defaultdict(lambda: {"footprint": "", "quantity": 0, "refs": []})

# Process components
for module in pcb.GetFootprints():
    if module.GetAttributes() & pcbnew.FP_EXCLUDE_FROM_BOM:
        continue  # Skip components marked as "Exclude from BOM"

    ref = module.GetReference()
    value = module.GetValue()
    footprint = str(module.GetFPID().GetLibItemName())

    tolerance = module.GetProperty("Tolerance") if module.HasProperty("Tolerance") else ""
    voltage = module.GetProperty("Voltage") if module.HasProperty("Voltage") else ""
    power = module.GetProperty("Power") if module.HasProperty("Power") else ""

    value_parts = [value]
    if tolerance:
        value_parts.append(f"{tolerance}")
    if voltage:
        value_parts.append(voltage if "V" in voltage else f"{voltage}V")
    if power:
        value_parts.append(power if "W" in power else f"{power}W")
    
    formatted_value = " - ".join(value_parts)
    key = (formatted_value, footprint)
    bom_dict[key]["footprint"] = footprint
    bom_dict[key]["quantity"] += 1
    bom_dict[key]["refs"].append(ref)

# Function to sort designators numerically
def sort_designators(designators):
    return sorted(designators, key=lambda x: (
        ''.join(filter(str.isalpha, x)),  # Extract letters
        int(''.join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else 0  # Extract numbers
    ))

# Write to CSV file
bom_file = os.path.join(BOM_PATH, f"{PROJECT_NAME}.csv")
with open(bom_file, "w") as f:
    f.write("Quantity, Value, Footprint, Designators\n")
    for (formatted_value, footprint), data in sorted(bom_dict.items(), key=lambda x: x[1]["refs"][0]):
        refs = ", ".join(sort_designators(data["refs"]))  # Sort designators numerically
        f.write(f"{refs}, {formatted_value}, {footprint}, {data['quantity']}\n")

print("BOM CSV export completed successfully!")

# Create Excel (XLSX) file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOM"

# Add headers
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
headers = ["Designator", "Value", "Footprint", "Quantity"]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = center_align

# Add data to the sheet
row_num = 2
for (formatted_value, footprint), data in sorted(bom_dict.items(), key=lambda x: x[1]["refs"][0]):
    designator = ", ".join(sort_designators(data["refs"]))  # Sort designators numerically
    quantity = data["quantity"]

    ws.append([designator, formatted_value, footprint, quantity])
    row_num += 1

# Apply formatting to columns
for col_num in range(1, 5):
    column = get_column_letter(col_num)
    if col_num == 1:  # Designator column width adjustment
        ws.column_dimensions[column].width = 20  # Adjust only the first column width
        # Allow line breaking for the Designator column
        for row in range(1, row_num):
            cell = ws[f"{column}{row}"]
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    else:
        # Adjust all other columns based on content (without line breaking)
        max_length = 0
        for row in range(1, row_num):
            cell = ws[f"{column}{row}"]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 300)  # Limit the column width to 250
        ws.column_dimensions[column].width = adjusted_width
        for row in range(1, row_num):
            cell = ws[f"{column}{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Set background color for other cells (Very light grey)
light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
for row in range(2, row_num + 1):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.fill = light_grey_fill
        cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                     right=openpyxl.styles.borders.Side(style='thin'),
                                                     top=openpyxl.styles.borders.Side(style='thin'),
                                                     bottom=openpyxl.styles.borders.Side(style='thin'))

# Apply thicker borders to the outer edges
for row in range(1, row_num + 1):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        if row == 1:
            cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                         right=openpyxl.styles.borders.Side(style='thin'),
                                                         top=openpyxl.styles.borders.Side(style='thin'),
                                                         bottom=openpyxl.styles.borders.Side(style='thin'))
        elif row == row_num:
            cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                         right=openpyxl.styles.borders.Side(style='thin'),
                                                         top=openpyxl.styles.borders.Side(style='thin'),
                                                         bottom=openpyxl.styles.borders.Side(style='thin'))
        else:
            cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                         right=openpyxl.styles.borders.Side(style='thin'),
                                                         top=openpyxl.styles.borders.Side(style='thin'),
                                                         bottom=openpyxl.styles.borders.Side(style='thin'))

# Remove blank row if exists
if row_num > 2:
    ws.delete_rows(row_num)

# Save the Excel file
xlsx_file = os.path.join(BOM_PATH, f"{PROJECT_NAME}.xlsx")
wb.save(xlsx_file)

print("BOM Excel export completed successfully!")

# Export schematic PDF
SCHEMATIC_PDF = os.path.join(PDF_PATH, f"{PROJECT_NAME}_schematic.pdf")
run_kicad_command(f'kicad-cli sch export pdf "{SCH_FILE}" -o "{SCHEMATIC_PDF}"', "Schematic PDF export")

# Export PCB layout PDF
PCB_PDF = os.path.join(PDF_PATH, f"{PROJECT_NAME}_layout.pdf")
run_kicad_command(f'kicad-cli pcb export pdf "{PCB_FILE}" --layers F.Cu,B.Cu,F.SilkS,B.SilkS -o "{PCB_PDF}"', "PCB layout PDF export")

# Export Gerber files (all used layers)
run_kicad_command(f'kicad-cli pcb export gerbers "{PCB_FILE}" -o "{GERBER_PATH}"', "Gerber files export")

# Export Drill files (for all holes)
run_kicad_command(f'kicad-cli pcb export drill --format gerber "{PCB_FILE}" -o "{GERBER_PATH}"', "Drill files export")

# Export
